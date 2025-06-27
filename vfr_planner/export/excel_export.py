"""
Export de plans de vol vers Excel
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from typing import List, Dict, Any


def export_to_excel(flight_data: Dict[str, Any], legs_data: List[Dict],
                   filename: str = "flight_plan.xlsx"):
    """
    Exporter un plan de vol vers Excel.

    :param flight_data: Dictionnaire contenant les données générales du vol,
                        par exemple immatriculation, type d'aéronef, vitesse,
                        consommation, départ, destination, date, pilote, etc.
    :type flight_data: Dict[str, Any]
    :param legs_data: Liste de dictionnaires représentant les segments du vol,
                      chaque dictionnaire contient les clés comme 'from', 'to',
                      'distance', 'true_course', 'fuel_leg', 'eta', etc.
    :type legs_data: List[Dict[str, Any]]
    :param filename: Nom du fichier Excel de sortie.
    :type filename: str
    :raises Exception: En cas d'erreur lors de la génération ou sauvegarde du fichier Excel.
    """
    try:
        # Créer workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan de Vol VFR"

        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="000080")
        data_font = Font(size=10)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Titre principal
        ws.merge_cells('A1:P1')
        ws['A1'] = "PLAN DE VOL VFR - VISUAL FLIGHT RULES FLIGHT PLAN"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # Sous-titre avec date de génération
        ws.merge_cells('A2:P2')
        ws['A2'] = f"Généré le {datetime.now().strftime('%Y-%m-%d à %H:%M')}"
        ws['A2'].alignment = center_align
        ws['A2'].font = Font(size=10, italic=True)

        # Section informations avion (colonne gauche)
        row = 4
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "INFORMATIONS DE L'AÉRONEF"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_align

        aircraft_info = [
            ("Immatriculation:", flight_data.get('aircraft_id', 'N/A')),
            ("Type d'aéronef:", flight_data.get('aircraft_type', 'N/A')),
            ("Vitesse vraie (TAS):", f"{flight_data.get('tas', 'N/A')} kn"),
            ("Consommation:", f"{flight_data.get('fuel_burn', 'N/A')} GPH"),
            ("Capacité carburant:", f"{flight_data.get('fuel_capacity', 'N/A')} gal"),
            ("Réserve requise:", f"{flight_data.get('reserve_fuel', 'N/A')} min"),
        ]

        for i, (label, value) in enumerate(aircraft_info):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            ws[f'A{row_num}'].font = Font(bold=True)
            ws[f'B{row_num}'] = value
            ws[f'A{row_num}'].border = border
            ws[f'B{row_num}'].border = border

        # Section informations vol (colonne droite)
        row = 4
        ws.merge_cells(f'I{row}:P{row}')
        ws[f'I{row}'] = "INFORMATIONS DE VOL"
        ws[f'I{row}'].font = header_font
        ws[f'I{row}'].fill = header_fill
        ws[f'I{row}'].alignment = center_align

        flight_info = [
            ("Aérodrome de départ:", flight_data.get('departure', 'N/A')),
            ("Aérodrome d'arrivée:", flight_data.get('destination', 'N/A')),
            ("Date de vol:", flight_data.get('date', 'N/A')),
            ("Heure de départ (ETD):", flight_data.get('etd', 'N/A')),
            ("Pilote commandant:", flight_data.get('pilot', 'N/A')),
            ("Briefing météo:", flight_data.get('weather_brief', 'N/A')),
        ]

        for i, (label, value) in enumerate(flight_info):
            row_num = row + 1 + i
            ws[f'I{row_num}'] = label
            ws[f'I{row_num}'].font = Font(bold=True)
            ws[f'J{row_num}'] = value
            ws[f'I{row_num}'].border = border
            ws[f'J{row_num}'].border = border

        # Table des legs de navigation
        nav_start_row = row + len(aircraft_info) + 3
        ws.merge_cells(f'A{nav_start_row}:P{nav_start_row}')
        ws[f'A{nav_start_row}'] = "JOURNAL DE NAVIGATION"
        ws[f'A{nav_start_row}'].font = header_font
        ws[f'A{nav_start_row}'].fill = header_fill
        ws[f'A{nav_start_row}'].alignment = center_align

        # En-têtes du tableau de navigation
        headers = [
            "Leg", "De", "À", "Dist\n(NM)", "Cap Vrai\n(°)", "Cap Mag\n(°)",
            "Vent Dir\n(°)", "Vent Vit\n(kn)", "Vit Sol\n(kn)", "Temps\n(min)",
            "Carb Leg\n(gal)", "Carb Tot\n(gal)", "ETA", "Remarques"
        ]

        header_row = nav_start_row + 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Données des legs
        total_distance = 0
        total_time = 0
        total_fuel = 0

        for i, leg in enumerate(legs_data):
            row_num = header_row + 1 + i
            row_data = [
                i + 1,
                leg.get('from', ''),
                leg.get('to', ''),
                f"{leg.get('distance', 0):.1f}",
                f"{leg.get('true_course', 0):.0f}",
                f"{leg.get('mag_heading', 0):.0f}",
                f"{leg.get('wind_dir', 0):.0f}",
                f"{leg.get('wind_speed', 0):.0f}",
                f"{leg.get('ground_speed', 0):.0f}",
                f"{leg.get('leg_time', 0):.0f}",
                f"{leg.get('fuel_leg', 0):.1f}",
                f"{leg.get('fuel_total', 0):.1f}",
                leg.get('eta', ''),
                leg.get('remarks', '')
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                cell.alignment = center_align
                cell.font = data_font

            total_distance += leg.get('distance', 0)
            total_time = leg.get('total_time', 0)
            total_fuel = leg.get('fuel_total', 0)

        # Ligne des totaux
        totals_row = header_row + len(legs_data) + 1
        ws[f'A{totals_row}'] = "TOTAUX:"
        ws[f'A{totals_row}'].font = Font(bold=True, size=12)
        ws[f'D{totals_row}'] = f"{total_distance:.1f}"
        ws[f'D{totals_row}'].font = Font(bold=True)
        ws[f'J{totals_row}'] = f"{total_time:.0f}"
        ws[f'J{totals_row}'].font = Font(bold=True)
        ws[f'L{totals_row}'] = f"{total_fuel:.1f}"
        ws[f'L{totals_row}'].font = Font(bold=True)

        # Section résumé et vérifications
        summary_row = totals_row + 3
        ws.merge_cells(f'A{summary_row}:P{summary_row}')
        ws[f'A{summary_row}'] = "RÉSUMÉ ET VÉRIFICATIONS"
        ws[f'A{summary_row}'].font = header_font
        ws[f'A{summary_row}'].fill = header_fill
        ws[f'A{summary_row}'].alignment = center_align

        # Calculs de sécurité
        reserve_fuel = flight_data.get('reserve_fuel', 45) * flight_data.get('fuel_burn', 7.5) / 60
        total_fuel_required = total_fuel + reserve_fuel
        fuel_capacity = flight_data.get('fuel_capacity', 0)

        summary_info = [
            ("Temps total de vol:", f"{total_time / 60:.1f} heures ({total_time:.0f} minutes)"),
            ("Distance totale:", f"{total_distance:.1f} milles nautiques"),
            ("Carburant de route:", f"{total_fuel:.1f} gallons"),
            ("Carburant de réserve:", f"{reserve_fuel:.1f} gallons"),
            ("Carburant total requis:", f"{total_fuel_required:.1f} gallons"),
            ("Capacité réservoir:", f"{fuel_capacity:.1f} gallons"),
            ("Marge de sécurité:",
             f"{fuel_capacity - total_fuel_required:.1f} gallons" if fuel_capacity > 0 else "À vérifier"),
        ]

        for i, (label, value) in enumerate(summary_info):
            row_num = summary_row + 1 + i
            ws[f'A{row_num}'] = label
            ws[f'A{row_num}'].font = Font(bold=True)
            ws[f'B{row_num}'] = value

            # Colorer en rouge si carburant insuffisant
            if "Marge de sécurité" in label and fuel_capacity > 0 and fuel_capacity < total_fuel_required:
                ws[f'B{row_num}'].font = Font(color="FF0000", bold=True)

        # Signatures et approbations
        sig_row = summary_row + len(summary_info) + 3
        ws[f'A{sig_row}'] = "Pilote commandant:"
        ws[f'F{sig_row}'] = "____________________"
        ws[f'A{sig_row + 1}'] = "Date et heure:"
        ws[f'F{sig_row + 1}'] = "____________________"

        # Ajuster la largeur des colonnes
        column_widths = [6, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Sauvegarder
        wb.save(filename)
        print(f"Plan de vol Excel sauvegardé: {filename}")

    except Exception as e:
        raise Exception(f"Erreur lors de la génération Excel: {e}")


def create_simple_excel_export(flight_data: Dict[str, Any], legs_data: List[Dict],
                              filename: str = "simple_flight_plan.xlsx"):
    """
    Créer un export Excel simplifié sans formatage complexe.

    :param flight_data: Données du vol.
    :type flight_data: Dict[str, Any]
    :param legs_data: Données des segments.
    :type legs_data: List[Dict]
    :param filename: Nom du fichier.
    :type filename: str
    :raises Exception: En cas d'erreur lors de l'export Excel.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan VFR"

        # En-tête simple
        ws['A1'] = "PLAN DE VOL VFR"
        ws['A3'] = f"Avion: {flight_data.get('aircraft_id', 'N/A')}"
        ws['A4'] = f"Pilote: {flight_data.get('pilot', 'N/A')}"
        ws['A5'] = f"Date: {flight_data.get('date', 'N/A')}"

        # Table des legs
        start_row = 7
        headers = ["Leg", "De", "À", "Distance", "Cap", "Vent", "VS", "Temps", "Carburant"]

        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)

        for i, leg in enumerate(legs_data):
            row = start_row + 1 + i
            ws.cell(row=row, column=1, value=i + 1)
            ws.cell(row=row, column=2, value=leg.get('from', ''))
            ws.cell(row=row, column=3, value=leg.get('to', ''))
            ws.cell(row=row, column=4, value=f"{leg.get('distance', 0):.1f} NM")
            ws.cell(row=row, column=5, value=f"{leg.get('mag_heading', 0):.0f}°")
            ws.cell(row=row, column=6, value=f"{leg.get('wind_dir', 0):.0f}°/{leg.get('wind_speed', 0):.0f}kn")
            ws.cell(row=row, column=7, value=f"{leg.get('ground_speed', 0):.0f} kn")
            ws.cell(row=row, column=8, value=f"{leg.get('leg_time', 0):.0f} min")
            ws.cell(row=row, column=9, value=f"{leg.get('fuel_leg', 0):.1f} gal")

        wb.save(filename)
        print(f"Plan Excel simple sauvegardé: {filename}")

    except Exception as e:
        raise Exception(f"Erreur export Excel simple: {e}")


# Fonctions utilitaires
def format_time(minutes: float) -> str:
    """
    Formater le temps en heures:minutes.

    :param minutes: Durée en minutes.
    :type minutes: float
    :return: Temps formaté en chaîne 'HH:MM'.
    :rtype: str
    """
    hours = int(minutes // 60)
    mins = int(minutes % 60)
    return f"{hours:02d}:{mins:02d}"


def format_coordinates(lat: float, lon: float) -> str:
    """
    Formater les coordonnées géographiques en degrés décimaux avec direction.

    :param lat: Latitude.
    :type lat: float
    :param lon: Longitude.
    :type lon: float
    :return: Coordonnées formatées, par exemple '48.8566°N, 2.3522°E'.
    :rtype: str
    """
    lat_dir = "N" if lat >= 0 else "S"
    lon_dir = "E" if lon >= 0 else "W"
    return f"{abs(lat):.4f}°{lat_dir}, {abs(lon):.4f}°{lon_dir}"


def add_flight_summary_sheet(wb: openpyxl.Workbook, flight_data: Dict[str, Any],
                             legs_data: List[Dict]) -> None:
    """
    Ajouter une feuille de résumé au workbook Excel.

    :param wb: Workbook Excel où ajouter la feuille.
    :type wb: openpyxl.Workbook
    :param flight_data: Données du vol.
    :type flight_data: Dict[str, Any]
    :param legs_data: Données des segments.
    :type legs_data: List[Dict]
    :return: None
    """
    ws = wb.create_sheet("Résumé")

    # Calculs totaux
    total_distance = sum(leg.get('distance', 0) for leg in legs_data)
    total_time = legs_data[-1].get('total_time', 0) if legs_data else 0
    total_fuel = legs_data[-1].get('fuel_total', 0) if legs_data else 0

    # Contenu du résumé
    summary_data = [
        ["RÉSUMÉ DU VOL", ""],
        ["", ""],
        ["Distance totale:", f"{total_distance:.1f} NM"],
        ["Temps de vol:", format_time(total_time)],
        ["Carburant requis:", f"{total_fuel:.1f} gal"],
        ["", ""],
        ["Départ:", flight_data.get('departure', 'N/A')],
        ["Arrivée:", flight_data.get('destination', 'N/A')],
        ["Nombre de segments:", str(len(legs_data))],
        ["", ""],
        ["Avion:", flight_data.get('aircraft_id', 'N/A')],
        ["Type:", flight_data.get('aircraft_type', 'N/A')],
        ["Pilote:", flight_data.get('pilot', 'N/A')],
    ]

    for i, (label, value) in enumerate(summary_data, 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)

    # Ajuster largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15


if __name__ == "__main__":
    # Test de la fonction
    test_flight_data = {
        'aircraft_id': 'C-FXYZ',
        'aircraft_type': 'C172',
        'tas': 110,
        'departure': 'CYUL',
        'destination': 'CYQB',
        'date': '2025-06-17',
        'etd': '10:00',
        'pilot': 'Test Pilot',
        'fuel_capacity': 40,
        'fuel_burn': 7.5,
        'reserve_fuel': 45
    }

    test_legs_data = [
        {
            'from': 'CYUL',
            'to': 'CYQB',
            'distance': 145.3,
            'true_course': 45,
            'mag_heading': 58,
            'wind_dir': 270,
            'wind_speed': 15,
            'ground_speed': 125,
            'leg_time': 70,
            'total_time': 70,
            'fuel_leg': 8.7,
            'fuel_total': 8.7,
            'eta': '11:10',
            'remarks': 'Route directe'
        }
    ]

    try:
        export_to_excel(test_flight_data, test_legs_data, "test_export.xlsx")
        print("✅ Test d'export Excel réussi!")
    except Exception as e:
        print(f"❌ Erreur de test: {e}")
