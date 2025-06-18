"""
Module d'exportation de plan de vol VFR en format Excel
Projet MGA802 - Outil de planification de vol VFR

Auteurs: Antoine Gingras, Matthew Meyer, Richard Nguekam, Gabriel Wong-Lapierre
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from typing import Optional


class FlightPlanExporter:
    """
    Classe pour exporter un itinéraire de vol en format Excel standardisé VFR.
    Génère un plan de vol similaire au format Jeppesen VFR Navigation Log.
    """

    def __init__(self, itinerary):
        """
        Initialise l'exportateur avec un objet Itinerary.

        Args:
            itinerary: Objet Itinerary contenant les waypoints et legs calculés
        """
        self.itinerary = itinerary
        self.aircraft_info = {}
        self.flight_info = {}

    def set_aircraft_info(self, aircraft_id: str, aircraft_type: str, tas: float,
                          fuel_capacity: float, burn_rate: float, **kwargs):
        """
        Définit les informations de l'aéronef.

        Args:
            aircraft_id: Immatriculation de l'aéronef (ex: C-GXYZ)
            aircraft_type: Type d'aéronef (ex: C172)
            tas: Vitesse vraie (True Airspeed) en noeuds
            fuel_capacity: Capacité de carburant en gallons
            burn_rate: Consommation en gallons/heure
            **kwargs: Autres informations optionnelles
        """
        self.aircraft_info = {
            'aircraft_id': aircraft_id,
            'aircraft_type': aircraft_type,
            'tas': tas,
            'fuel_capacity': fuel_capacity,
            'burn_rate': burn_rate,
            **kwargs
        }

    def set_flight_info(self, pilot_name: str, departure_time: datetime.datetime,
                        fuel_on_board: float, passengers: int = 1, **kwargs):
        """
        Définit les informations du vol.

        Args:
            pilot_name: Nom du pilote
            departure_time: Heure de départ
            fuel_on_board: Carburant à bord en gallons
            passengers: Nombre de personnes à bord
            **kwargs: Autres informations optionnelles
        """
        self.flight_info = {
            'pilot_name': pilot_name,
            'departure_time': departure_time,
            'fuel_on_board': fuel_on_board,
            'passengers': passengers,
            **kwargs
        }

    def _create_header_section(self, ws):
        """
        Crée la section d'en-tête du plan de vol avec les informations générales.
        """
        # Titre principal
        ws['A1'] = 'VFR NAVIGATION LOG'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:O1')

        # Informations aéronef (colonne gauche)
        ws['A3'] = 'Aircraft Number:'
        ws['B3'] = self.aircraft_info.get('aircraft_id', '')
        ws['A4'] = 'Aircraft Type:'
        ws['B4'] = self.aircraft_info.get('aircraft_type', '')
        ws['A5'] = 'TAS:'
        ws['B5'] = f"{self.aircraft_info.get('tas', 0)} kts"

        # Informations vol (colonne droite)
        ws['H3'] = 'Pilot:'
        ws['I3'] = self.flight_info.get('pilot_name', '')
        ws['H4'] = 'Departure Time:'
        departure_time = self.flight_info.get('departure_time')
        if departure_time:
            ws['I4'] = departure_time.strftime('%H:%M Z')
        ws['H5'] = 'Fuel on Board:'
        ws['I5'] = f"{self.flight_info.get('fuel_on_board', 0):.1f} gal"
        ws['H6'] = 'Persons on Board:'
        ws['I6'] = self.flight_info.get('passengers', 1)

        return 8  # Retourne la prochaine ligne disponible

    def _create_navigation_table(self, ws, start_row):
        """
        Crée le tableau principal de navigation avec tous les legs du vol.
        """
        # En-têtes du tableau
        headers = [
            'Check Points\n(Fixes)', 'Course\n(TC)', 'Distance\n(NM)',
            'Wind\nDir/Vel', 'WCA\n(°)', 'TH\n(°)', 'MH\n(°)', 'GS\n(kts)',
            'ETE\n(min)', 'ATE\n(min)', 'Fuel Used\n(gal)', 'Fuel Rem.\n(gal)',
            'Altitude\n(ft)', 'Remarks'
        ]

        # Appliquer les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Appliquer les bordures aux en-têtes
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col in range(1, len(headers) + 1):
            ws.cell(row=start_row, column=col).border = thin_border

        # Créer les legs si pas déjà fait
        if not self.itinerary.legs:
            self.itinerary.create_legs()

        # Remplir les données des legs
        current_row = start_row + 1
        total_fuel_used = 0
        fuel_remaining = self.flight_info.get('fuel_on_board', 0)

        for i, leg in enumerate(self.itinerary.legs):
            # Calculer le carburant restant
            total_fuel_used += leg.fuel_burn_leg
            fuel_remaining = self.flight_info.get('fuel_on_board', 0) - total_fuel_used

            # Remplir les données de la ligne
            row_data = [
                f"{leg.starting_wp.name} → {leg.ending_wp.name}",  # Check Points
                f"{leg.tc:.0f}°",  # True Course
                f"{leg.distance:.1f}",  # Distance
                f"{leg.wind_dir:.0f}°/{leg.wind_speed:.0f}",  # Wind
                f"{leg.wca:.1f}°",  # Wind Correction Angle
                f"{leg.th:.0f}°",  # True Heading
                f"{leg.mh:.0f}°",  # Magnetic Heading
                f"{leg.gs:.0f}",  # Ground Speed
                f"{leg.time_leg:.0f}",  # Estimated Time Enroute
                f"{leg.time_tot:.0f}",  # Accumulated Time Enroute
                f"{leg.fuel_burn_leg:.1f}",  # Fuel Used this leg
                f"{fuel_remaining:.1f}",  # Fuel Remaining
                f"{leg.starting_wp.alt or 0:.0f}",  # Altitude
                ""  # Remarks
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            current_row += 1

        # Ajouter une ligne de totaux
        ws.cell(row=current_row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=current_row, column=3, value=f"{sum(leg.distance for leg in self.itinerary.legs):.1f}")
        ws.cell(row=current_row, column=9, value=f"{self.itinerary.legs[-1].time_tot:.0f}")
        ws.cell(row=current_row, column=11, value=f"{total_fuel_used:.1f}")

        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).font = Font(bold=True)

        return current_row + 2

    def _create_weather_section(self, ws, start_row):
        """
        Crée une section pour les informations météorologiques.
        """
        ws.cell(row=start_row, column=1, value="WEATHER INFORMATION").font = Font(bold=True, size=12)

        start_row += 2
        weather_headers = ['Time', 'Wind Direction', 'Wind Speed', 'Temperature', 'Visibility', 'Remarks']

        for col, header in enumerate(weather_headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

        # Ajouter quelques lignes vides pour remplissage manuel
        for row in range(start_row + 1, start_row + 4):
            for col in range(1, len(weather_headers) + 1):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        return start_row + 5

    def _format_worksheet(self, ws):
        """
        Applique le formatage général à la feuille de calcul.
        """
        # Ajuster la largeur des colonnes
        column_widths = [20, 10, 10, 12, 8, 8, 8, 10, 10, 10, 12, 12, 10, 15]

        for i, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[column_letter].width = width

        # Ajuster la hauteur des lignes d'en-tête
        ws.row_dimensions[8].height = 40  # En-têtes du tableau principal

    def export_to_excel(self, filename: str, include_weather: bool = True):
        """
        Exporte le plan de vol vers un fichier Excel.

        Args:
            filename: Nom du fichier de sortie (avec extension .xlsx)
            include_weather: Inclure la section météo (défaut: True)
        """
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VFR Flight Plan"

        # Créer les différentes sections
        current_row = self._create_header_section(ws)
        current_row = self._create_navigation_table(ws, current_row)

        if include_weather:
            current_row = self._create_weather_section(ws, current_row)

        # Appliquer le formatage
        self._format_worksheet(ws)

        # Sauvegarder le fichier
        wb.save(filename)
        print(f"Plan de vol exporté vers: {filename}")

    def export_summary_dataframe(self) -> pd.DataFrame:
        """
        Retourne un DataFrame pandas avec un résumé du plan de vol.

        Returns:
            DataFrame contenant les informations essentielles du vol
        """
        if not self.itinerary.legs:
            self.itinerary.create_legs()

        data = []
        for i, leg in enumerate(self.itinerary.legs):
            data.append({
                'Leg': i + 1,
                'From': leg.starting_wp.name,
                'To': leg.ending_wp.name,
                'Distance_NM': round(leg.distance, 1),
                'True_Course': round(leg.tc, 0),
                'True_Heading': round(leg.th, 0),
                'Magnetic_Heading': round(leg.mh, 0),
                'Ground_Speed': round(leg.gs, 0),
                'ETE_minutes': round(leg.time_leg, 0),
                'Cumulative_Time': round(leg.time_tot, 0),
                'Fuel_Used_gal': round(leg.fuel_burn_leg, 1),
                'Wind_Direction': round(leg.wind_dir, 0),
                'Wind_Speed': round(leg.wind_speed, 0)
            })

        return pd.DataFrame(data)


# Fonction utilitaire pour faciliter l'utilisation
def export_flight_plan(itinerary, filename: str, aircraft_info: dict, flight_info: dict):
    """
    Fonction utilitaire pour exporter rapidement un plan de vol.

    Args:
        itinerary: Objet Itinerary avec les waypoints et legs
        filename: Nom du fichier Excel de sortie
        aircraft_info: Dict avec les infos aéronef (aircraft_id, aircraft_type, tas, etc.)
        flight_info: Dict avec les infos vol (pilot_name, departure_time, etc.)
    """
    exporter = FlightPlanExporter(itinerary)
    exporter.set_aircraft_info(**aircraft_info)
    exporter.set_flight_info(**flight_info)
    exporter.export_to_excel(filename)

    return exporter.export_summary_dataframe()


# Exemple d'utilisation
if __name__ == "__main__":
    # Cet exemple montre comment utiliser le module
    # (nécessiterait un objet Itinerary réel pour fonctionner)

    print("Module d'exportation de plan de vol VFR")
    print("Usage:")
    print("1. Créer un objet Itinerary avec vos waypoints")
    print("2. Créer un FlightPlanExporter(itinerary)")
    print("3. Définir les infos aéronef avec set_aircraft_info()")
    print("4. Définir les infos vol avec set_flight_info()")
    print("5. Exporter avec export_to_excel(filename)")