import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict, Any


class FlightPlanGenerator:
    def __init__(self):
        self.styles = getSampleStyleSheet()

    def generate_excel_plan(self, flight_data: Dict[str, Any], legs_data: List[Dict],
                            filename: str = "flight_plan.xlsx"):
        """Générer un plan de vol en format Excel"""

        # Créer workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan de Vol VFR"

        # Styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Titre principal
        ws.merge_cells('A1:P1')
        ws['A1'] = "PLAN DE VOL VFR - VISUAL FLIGHT RULES FLIGHT PLAN"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # Informations générales
        row = 3
        general_info = [
            ("Aircraft ID:", flight_data.get('aircraft_id', '')),
            ("Aircraft Type:", flight_data.get('aircraft_type', '')),
            ("True Airspeed:", f"{flight_data.get('tas', '')} kn"),
            ("Departure:", flight_data.get('departure', '')),
            ("Destination:", flight_data.get('destination', '')),
            ("Date:", flight_data.get('date', '')),
            ("ETD:", flight_data.get('etd', '')),
            ("Pilot:", flight_data.get('pilot', '')),
            ("Fuel Capacity:", f"{flight_data.get('fuel_capacity', '')} gal"),
            ("Fuel Burn Rate:", f"{flight_data.get('fuel_burn', '')} GPH"),
        ]

        for i, (label, value) in enumerate(general_info[:5]):
            ws.cell(row=row + i, column=1, value=label).font = header_font
            ws.cell(row=row + i, column=2, value=value)

        for i, (label, value) in enumerate(general_info[5:]):
            ws.cell(row=row + i, column=9, value=label).font = header_font
            ws.cell(row=row + i, column=10, value=value)

        # Table des legs
        row = 14
        headers = [
            "Leg", "From", "To", "Distance (NM)", "True Course", "True Heading",
            "Mag Heading", "Wind Dir", "Wind Speed", "Ground Speed",
            "Leg Time (min)", "Total Time", "Fuel Burn Leg", "Fuel Total", "ETA", "Remarks"
        ]

        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Données des legs
        total_distance = 0
        total_time = 0
        total_fuel = 0

        for i, leg in enumerate(legs_data, 1):
            row_data = [
                i,
                leg.get('from', ''),
                leg.get('to', ''),
                f"{leg.get('distance', 0):.1f}",
                f"{leg.get('true_course', 0):.0f}°",
                f"{leg.get('true_heading', 0):.0f}°",
                f"{leg.get('mag_heading', 0):.0f}°",
                f"{leg.get('wind_dir', 0):.0f}°",
                f"{leg.get('wind_speed', 0):.0f} kn",
                f"{leg.get('ground_speed', 0):.0f} kn",
                f"{leg.get('leg_time', 0):.0f}",
                f"{leg.get('total_time', 0):.0f}",
                f"{leg.get('fuel_leg', 0):.1f} gal",
                f"{leg.get('fuel_total', 0):.1f} gal",
                leg.get('eta', ''),
                leg.get('remarks', '')
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row + i, column=col, value=value)
                cell.border = border
                cell.alignment = center_align

            total_distance += leg.get('distance', 0)
            total_time = leg.get('total_time', 0)
            total_fuel = leg.get('fuel_total', 0)

        # Totaux
        totals_row = row + len(legs_data) + 1
        ws.cell(row=totals_row, column=1, value="TOTALS:").font = header_font
        ws.cell(row=totals_row, column=4, value=f"{total_distance:.1f} NM").font = header_font
        ws.cell(row=totals_row, column=12, value=f"{total_time:.0f} min").font = header_font
        ws.cell(row=totals_row, column=14, value=f"{total_fuel:.1f} gal").font = header_font

        # Informations supplémentaires
        additional_info_row = totals_row + 3
        additional_info = [
            ("Reserve Fuel:", f"{flight_data.get('reserve_fuel', 45)} min"),
            ("Alternate Airport:", flight_data.get('alternate', '')),
            ("Weather Brief:", flight_data.get('weather_brief', '')),
            ("NOTAM Check:", flight_data.get('notam_check', '')),
            ("Flight Following:", flight_data.get('flight_following', '')),
        ]

        for i, (label, value) in enumerate(additional_info):
            ws.cell(row=additional_info_row + i, column=1, value=label).font = header_font
            ws.cell(row=additional_info_row + i, column=2, value=value)

        # Ajuster la largeur des colonnes
        column_widths = [8, 12, 12, 12, 12, 12, 12, 10, 12, 12, 12, 12, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Sauvegarder
        wb.save(filename)
        print(f"Plan de vol Excel sauvegardé: {filename}")

    def generate_pdf_plan(self, flight_data: Dict[str, Any], legs_data: List[Dict],
                          filename: str = "flight_plan.pdf"):
        """Générer un plan de vol en format PDF"""

        doc = SimpleDocTemplate(filename, pagesize=letter,
                                leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                                topMargin=0.5 * inch, bottomMargin=0.5 * inch)

        story = []

        # Titre
        title = Paragraph("PLAN DE VOL VFR - VISUAL FLIGHT RULES FLIGHT PLAN",
                          self.styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))

        # Informations générales en tableau
        general_data = [
            ["Aircraft ID:", flight_data.get('aircraft_id', ''),
             "Date:", flight_data.get('date', '')],
            ["Aircraft Type:", flight_data.get('aircraft_type', ''),
             "ETD:", flight_data.get('etd', '')],
            ["True Airspeed:", f"{flight_data.get('tas', '')} kn",
             "Pilot:", flight_data.get('pilot', '')],
            ["Departure:", flight_data.get('departure', ''),
             "Fuel Capacity:", f"{flight_data.get('fuel_capacity', '')} gal"],
            ["Destination:", flight_data.get('destination', ''),
             "Fuel Burn:", f"{flight_data.get('fuel_burn', '')} GPH"],
        ]

        general_table = Table(general_data, colWidths=[1.5 * inch, 2 * inch, 1.5 * inch, 2 * inch])
        general_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(general_table)
        story.append(Spacer(1, 20))

        # Table des legs
        headers = [
            "Leg", "From", "To", "Dist\n(NM)", "TC\n(°)", "TH\n(°)", "MH\n(°)",
            "Wind\nDir", "Wind\nSpd", "GS\n(kn)", "Time\n(min)", "Fuel\n(gal)", "ETA"
        ]

        legs_table_data = [headers]

        for i, leg in enumerate(legs_data, 1):
            row = [
                str(i),
                leg.get('from', '')[:6],  # Limiter la longueur
                leg.get('to', '')[:6],
                f"{leg.get('distance', 0):.0f}",
                f"{leg.get('true_course', 0):.0f}",
                f"{leg.get('true_heading', 0):.0f}",
                f"{leg.get('mag_heading', 0):.0f}",
                f"{leg.get('wind_dir', 0):.0f}",
                f"{leg.get('wind_speed', 0):.0f}",
                f"{leg.get('ground_speed', 0):.0f}",
                f"{leg.get('leg_time', 0):.0f}",
                f"{leg.get('fuel_leg', 0):.1f}",
                leg.get('eta', '')[:5]
            ]
            legs_table_data.append(row)

        # Calculer largeurs colonnes proportionnellement
        col_widths = [0.4, 0.7, 0.7, 0.5, 0.4, 0.4, 0.4, 0.5, 0.5, 0.5, 0.6, 0.6, 0.6]
        col_widths = [w * inch for w in col_widths]

        legs_table = Table(legs_table_data, colWidths=col_widths, repeatRows=1)
        legs_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(Paragraph("NAVIGATION LOG", self.styles['Heading2']))
        story.append(legs_table)
        story.append(Spacer(1, 20))

        # Informations supplémentaires
        additional_info = f"""
        <b>ADDITIONAL INFORMATION:</b><br/>
        Reserve Fuel: {flight_data.get('reserve_fuel', 45)} minutes<br/>
        Alternate Airport: {flight_data.get('alternate', 'N/A')}<br/>
        Weather Briefing: {flight_data.get('weather_brief', 'N/A')}<br/>
        NOTAM Check: {flight_data.get('notam_check', 'N/A')}<br/>
        Flight Following: {flight_data.get('flight_following', 'N/A')}<br/>
        <br/>
        <b>Pilot Signature:</b> ___________________ <b>Date:</b> ___________
        """

        story.append(Paragraph(additional_info, self.styles['Normal']))

        # Générer PDF
        doc.build(story)
        print(f"Plan de vol PDF sauvegardé: {filename}")


# Exemple d'utilisation
def example_usage():
    # Données d'exemple
    flight_data = {
        'aircraft_id': 'C-FXYZ',
        'aircraft_type': 'C172',
        'tas': 110,
        'departure': 'CYUL',
        'destination': 'CYQB',
        'date': '2025-06-17',
        'etd': '10:00',
        'pilot': 'Jean Dupont',
        'fuel_capacity': 40,
        'fuel_burn': 7.5,
        'reserve_fuel': 45,
        'alternate': 'CYTR',
        'weather_brief': 'Obtained 09:30Z',
        'notam_check': 'Completed',
        'flight_following': 'Requested'
    }

    legs_data = [
        {
            'from': 'CYUL',
            'to': 'CYQB',
            'distance': 145.3,
            'true_course': 045,
            'true_heading': 043,
            'mag_heading': 058,
        'wind_dir': 270,
    'wind_speed': 15,
    'ground_speed': 125,
    'leg_time': 70,
    'total_time': 70,
    'fuel_leg': 8.7,
    'fuel_total': 8.7,
    'eta': '11:10',
    'remarks': ''
    }
    ]

    generator = FlightPlanGenerator()
    generator.generate_excel_plan(flight_data, legs_data, "example_flight_plan.xlsx")
    generator.generate_pdf_plan(flight_data, legs_data, "example_flight_plan.pdf")


if __name__ == "__main__":
    example_usage()