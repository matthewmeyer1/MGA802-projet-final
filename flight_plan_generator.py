# flight_plan_generator.py
"""
Générateur de plans de vol VFR en formats Excel et PDF
Intégré avec l'interface GUI du projet MGA802
"""

import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from typing import List, Dict, Any


class FlightPlanGenerator:
    """Générateur de plans de vol VFR professionnels"""

    def __init__(self):
        self.styles = getSampleStyleSheet()

    def generate_excel_plan(self, flight_data: Dict[str, Any], legs_data: List[Dict],
                            filename: str = "flight_plan.xlsx"):
        """Générer un plan de vol en format Excel professionnel"""

        try:
            # Créer workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Plan de Vol VFR"

            # Styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            title_font = Font(bold=True, size=16, color="000080")
            data_font = Font(size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Titre principal
            ws.merge_cells('A1:P1')
            ws['A1'] = "PLAN DE VOL VFR - VISUAL FLIGHT RULES FLIGHT PLAN"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align

            # Sous-titre avec date de génération
            ws.merge_cells('A2:P2')
            ws['A2'] = f"Généré le {datetime.now().strftime('%Y-%m-%d à %H:%M')}"
            ws['A2'].alignment = center_align
            ws['A2'].font = Font(size=10, italic=True)

            # Section informations avion (colonne gauche)
            row = 4
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = "INFORMATIONS DE L'AÉRONEF"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_align

            aircraft_info = [
                ("Immatriculation:", flight_data.get('aircraft_id', 'N/A')),
                ("Type d'aéronef:", flight_data.get('aircraft_type', 'N/A')),
                ("Vitesse vraie (TAS):", f"{flight_data.get('tas', 'N/A')} kn"),
                ("Consommation:", f"{flight_data.get('fuel_burn', 'N/A')} GPH"),
                ("Capacité carburant:", f"{flight_data.get('fuel_capacity', 'N/A')} gal"),
                ("Réserve requise:", f"{flight_data.get('reserve_fuel', 'N/A')} min"),
            ]

            for i, (label, value) in enumerate(aircraft_info):
                row_num = row + 1 + i
                ws[f'A{row_num}'] = label
                ws[f'A{row_num}'].font = Font(bold=True)
                ws[f'B{row_num}'] = value
                ws[f'A{row_num}'].border = border
                ws[f'B{row_num}'].border = border

            # Section informations vol (colonne droite)
            row = 4
            ws.merge_cells(f'I{row}:P{row}')
            ws[f'I{row}'] = "INFORMATIONS DE VOL"
            ws[f'I{row}'].font = header_font
            ws[f'I{row}'].fill = header_fill
            ws[f'I{row}'].alignment = center_align

            flight_info = [
                ("Aérodrome de départ:", flight_data.get('departure', 'N/A')),
                ("Aérodrome d'arrivée:", flight_data.get('destination', 'N/A')),
                ("Date de vol:", flight_data.get('date', 'N/A')),
                ("Heure de départ (ETD):", flight_data.get('etd', 'N/A')),
                ("Pilote commandant:", flight_data.get('pilot', 'N/A')),
                ("Briefing météo:", flight_data.get('weather_brief', 'N/A')),
            ]

            for i, (label, value) in enumerate(flight_info):
                row_num = row + 1 + i
                ws[f'I{row_num}'] = label
                ws[f'I{row_num}'].font = Font(bold=True)
                ws[f'J{row_num}'] = value
                ws[f'I{row_num}'].border = border
                ws[f'J{row_num}'].border = border

            # Table des legs de navigation
            nav_start_row = row + len(aircraft_info) + 3
            ws.merge_cells(f'A{nav_start_row}:P{nav_start_row}')
            ws[f'A{nav_start_row}'] = "JOURNAL DE NAVIGATION"
            ws[f'A{nav_start_row}'].font = header_font
            ws[f'A{nav_start_row}'].fill = header_fill
            ws[f'A{nav_start_row}'].alignment = center_align

            # En-têtes du tableau de navigation
            headers = [
                "Leg", "De", "À", "Dist\n(NM)", "Cap Vrai\n(°)", "Cap Mag\n(°)",
                "Vent Dir\n(°)", "Vent Vit\n(kn)", "Vit Sol\n(kn)", "Temps\n(min)",
                "Carb Leg\n(gal)", "Carb Tot\n(gal)", "ETA", "Remarques"
            ]

            header_row = nav_start_row + 1
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # Données des legs
            total_distance = 0
            total_time = 0
            total_fuel = 0

            for i, leg in enumerate(legs_data):
                row_num = header_row + 1 + i
                row_data = [
                    i + 1,
                    leg.get('from', ''),
                    leg.get('to', ''),
                    f"{leg.get('distance', 0):.1f}",
                    f"{leg.get('true_course', 0):.0f}",
                    f"{leg.get('mag_heading', 0):.0f}",
                    f"{leg.get('wind_dir', 0):.0f}",
                    f"{leg.get('wind_speed', 0):.0f}",
                    f"{leg.get('ground_speed', 0):.0f}",
                    f"{leg.get('leg_time', 0):.0f}",
                    f"{leg.get('fuel_leg', 0):.1f}",
                    f"{leg.get('fuel_total', 0):.1f}",
                    leg.get('eta', ''),
                    leg.get('remarks', '')
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border
                    cell.alignment = center_align
                    cell.font = data_font

                total_distance += leg.get('distance', 0)
                total_time = leg.get('total_time', 0)  # Temps cumulé du dernier leg
                total_fuel = leg.get('fuel_total', 0)  # Carburant cumulé du dernier leg

            # Ligne des totaux
            totals_row = header_row + len(legs_data) + 1
            ws[f'A{totals_row}'] = "TOTAUX:"
            ws[f'A{totals_row}'].font = Font(bold=True, size=12)
            ws[f'D{totals_row}'] = f"{total_distance:.1f}"
            ws[f'D{totals_row}'].font = Font(bold=True)
            ws[f'J{totals_row}'] = f"{total_time:.0f}"
            ws[f'J{totals_row}'].font = Font(bold=True)
            ws[f'L{totals_row}'] = f"{total_fuel:.1f}"
            ws[f'L{totals_row}'].font = Font(bold=True)

            # Section résumé et vérifications
            summary_row = totals_row + 3
            ws.merge_cells(f'A{summary_row}:P{summary_row}')
            ws[f'A{summary_row}'] = "RÉSUMÉ ET VÉRIFICATIONS"
            ws[f'A{summary_row}'].font = header_font
            ws[f'A{summary_row}'].fill = header_fill
            ws[f'A{summary_row}'].alignment = center_align

            # Calculs de sécurité
            reserve_fuel = flight_data.get('reserve_fuel', 45) * flight_data.get('fuel_burn', 7.5) / 60
            total_fuel_required = total_fuel + reserve_fuel
            fuel_capacity = flight_data.get('fuel_capacity', 0)

            summary_info = [
                ("Temps total de vol:", f"{total_time / 60:.1f} heures ({total_time:.0f} minutes)"),
                ("Distance totale:", f"{total_distance:.1f} milles nautiques"),
                ("Carburant de route:", f"{total_fuel:.1f} gallons"),
                ("Carburant de réserve:", f"{reserve_fuel:.1f} gallons"),
                ("Carburant total requis:", f"{total_fuel_required:.1f} gallons"),
                ("Capacité réservoir:", f"{fuel_capacity:.1f} gallons"),
                ("Marge de sécurité:",
                 f"{fuel_capacity - total_fuel_required:.1f} gallons" if fuel_capacity > 0 else "À vérifier"),
            ]

            for i, (label, value) in enumerate(summary_info):
                row_num = summary_row + 1 + i
                ws[f'A{row_num}'] = label
                ws[f'A{row_num}'].font = Font(bold=True)
                ws[f'B{row_num}'] = value

                # Colorer en rouge si carburant insuffisant
                if "Marge de sécurité" in label and fuel_capacity > 0 and fuel_capacity < total_fuel_required:
                    ws[f'B{row_num}'].font = Font(color="FF0000", bold=True)

            # Signatures et approbations
            sig_row = summary_row + len(summary_info) + 3
            ws[f'A{sig_row}'] = "Pilote commandant:"
            ws[f'F{sig_row}'] = "____________________"
            ws[f'A{sig_row + 1}'] = "Date et heure:"
            ws[f'F{sig_row + 1}'] = "____________________"

            # Ajuster la largeur des colonnes
            column_widths = [6, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            # Sauvegarder
            wb.save(filename)
            print(f"Plan de vol Excel sauvegardé: {filename}")

        except Exception as e:
            raise Exception(f"Erreur lors de la génération Excel: {e}")

    def generate_pdf_plan(self, flight_data: Dict[str, Any], legs_data: List[Dict],
                          filename: str = "flight_plan.pdf"):
        """Générer un plan de vol en format PDF professionnel"""

        try:
            doc = SimpleDocTemplate(filename, pagesize=letter,
                                    leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                                    topMargin=0.5 * inch, bottomMargin=0.5 * inch)

            story = []

            # Titre
            title = Paragraph("PLAN DE VOL VFR - VISUAL FLIGHT RULES FLIGHT PLAN",
                              self.styles['Title'])
            story.append(title)

            subtitle = Paragraph(f"Généré le {datetime.now().strftime('%Y-%m-%d à %H:%M')}",
                                 self.styles['Normal'])
            story.append(subtitle)
            story.append(Spacer(1, 12))

            # Informations générales en tableau
            general_data = [
                ["AÉRONEF", "", "VOL", ""],
                ["Immatriculation:", flight_data.get('aircraft_id', 'N/A'),
                 "Départ:", flight_data.get('departure', 'N/A')],
                ["Type:", flight_data.get('aircraft_type', 'N/A'),
                 "Destination:", flight_data.get('destination', 'N/A')],
                ["TAS:", f"{flight_data.get('tas', 'N/A')} kn",
                 "Date:", flight_data.get('date', 'N/A')],
                ["Consommation:", f"{flight_data.get('fuel_burn', 'N/A')} GPH",
                 "ETD:", flight_data.get('etd', 'N/A')],
                ["Capacité:", f"{flight_data.get('fuel_capacity', 'N/A')} gal",
                 "Pilote:", flight_data.get('pilot', 'N/A')],
            ]

            general_table = Table(general_data, colWidths=[1.2 * inch, 1.8 * inch, 1.2 * inch, 1.8 * inch])
            general_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.darkblue),
                ('BACKGROUND', (2, 0), (3, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (3, 0), colors.whitesmoke),
                ('BACKGROUND', (0, 1), (0, -1), colors.lightgrey),
                ('BACKGROUND', (2, 1), (2, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(general_table)
            story.append(Spacer(1, 20))

            # Table des legs
            headers = [
                "Leg", "De", "À", "Dist\n(NM)", "CV\n(°)", "CM\n(°)", "Vent\nDir", "Vent\nVit",
                "VS\n(kn)", "Temps\n(min)", "Carb\n(gal)", "ETA"
            ]

            legs_table_data = [headers]

            for i, leg in enumerate(legs_data, 1):
                row = [
                    str(i),
                    leg.get('from', '')[:5],
                    leg.get('to', '')[:5],
                    f"{leg.get('distance', 0):.0f}",
                    f"{leg.get('true_course', 0):.0f}",
                    f"{leg.get('mag_heading', 0):.0f}",
                    f"{leg.get('wind_dir', 0):.0f}",
                    f"{leg.get('wind_speed', 0):.0f}",
                    f"{leg.get('ground_speed', 0):.0f}",
                    f"{leg.get('leg_time', 0):.0f}",
                    f"{leg.get('fuel_leg', 0):.1f}",
                    leg.get('eta', '')[:5]
                ]
                legs_table_data.append(row)

            # Calculer largeurs colonnes
            col_widths = [0.4, 0.6, 0.6, 0.5, 0.4, 0.4, 0.4, 0.4, 0.4, 0.5, 0.5, 0.5]
            col_widths = [w * inch for w in col_widths]

            legs_table = Table(legs_table_data, colWidths=col_widths, repeatRows=1)
            legs_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))

            story.append(Paragraph("JOURNAL DE NAVIGATION", self.styles['Heading2']))
            story.append(legs_table)
            story.append(Spacer(1, 20))

            # Calculs et vérifications
            total_distance = sum(leg.get('distance', 0) for leg in legs_data)
            total_time = legs_data[-1].get('total_time', 0) if legs_data else 0
            total_fuel = legs_data[-1].get('fuel_total', 0) if legs_data else 0
            reserve_fuel = flight_data.get('reserve_fuel', 45) * flight_data.get('fuel_burn', 7.5) / 60
            total_fuel_required = total_fuel + reserve_fuel

            calculations = f"""
            <b>CALCULS ET VÉRIFICATIONS:</b><br/>
            <br/>
            Distance totale: <b>{total_distance:.1f} milles nautiques</b><br/>
            Temps total de vol: <b>{total_time / 60:.1f} heures ({total_time:.0f} minutes)</b><br/>
            Carburant de route: <b>{total_fuel:.1f} gallons</b><br/>
            Carburant de réserve: <b>{reserve_fuel:.1f} gallons</b><br/>
            <b>Carburant total requis: {total_fuel_required:.1f} gallons</b><br/>
            <br/>
            <b>BRIEFINGS ET VÉRIFICATIONS:</b><br/>
            Briefing météo: {flight_data.get('weather_brief', 'À compléter')}<br/>
            Vérification NOTAM: {flight_data.get('notam_check', 'Requise')}<br/>
            Suivi de vol: {flight_data.get('flight_following', 'Recommandé')}<br/>
            <br/>
            <b>SIGNATURES:</b><br/>
            <br/>
            Pilote commandant: _________________________ Date: __________<br/>
            <br/>
            Instructeur (si requis): _____________________ Date: __________<br/>
            """

            story.append(Paragraph(calculations, self.styles['Normal']))

            # Générer PDF
            doc.build(story)
            print(f"Plan de vol PDF sauvegardé: {filename}")

        except Exception as e:
            raise Exception(f"Erreur lors de la génération PDF: {e}")


# Test du générateur
if __name__ == "__main__":
    # Données d'exemple pour test
    test_flight_data = {
        'aircraft_id': 'C-FXYZ',
        'aircraft_type': 'C172',
        'tas': 110,
        'departure': 'CYUL',
        'destination': 'CYQB',
        'date': '2025-06-17',
        'etd': '10:00',
        'pilot': 'Test Pilot',
        'fuel_capacity': 40,
        'fuel_burn': 7.5,
        'reserve_fuel': 45,
        'weather_brief': 'Obtained via Tomorrow.io API',
        'notam_check': 'Completed',
        'flight_following': 'Requested'
    }

    test_legs_data = [
        {
            'from': 'CYUL',
            'to': 'CYQB',
            'distance': 145.3,
            'true_course': 45,
            'mag_heading': 58,
            'wind_dir': 270,
            'wind_speed': 15,
            'ground_speed': 125,
            'leg_time': 70,
            'total_time': 70,
            'fuel_leg': 8.7,
            'fuel_total': 8.7,
            'eta': '11:10',
            'remarks': 'Direct route'
        }
    ]

    generator = FlightPlanGenerator()
    try:
        generator.generate_excel_plan(test_flight_data, test_legs_data, "test_plan.xlsx")
        generator.generate_pdf_plan(test_flight_data, test_legs_data, "test_plan.pdf")
        print("✅ Tests de génération réussis!")
    except Exception as e:
        print(f"❌ Erreur de test: {e}")